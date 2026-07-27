"""
Compliance & Reporting API
"""
import io
import json
import re
from datetime import datetime, timedelta, timezone
from typing import List
from uuid import UUID
from xml.sax.saxutils import escape

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.core.deps import get_db, get_current_user
from app.db.models.user import User
from app.db.models.compliance import AuditLog, IncidentReport
from app.db.models.threat import Alert, Threat
from app.db.models.login_history import LoginHistory
from app.db.models.simulation import AttackSimulation
from app.db.models.security import SecurityScore
from app.db.models.session import UserSession
from app.db.models.device import Device
from app.db.models.identity_security import PasskeyCredential, NotificationIntegration
from app.schemas.common import AuditLogOut, IncidentReportOut
from app.services.llm_service import generate_executive_report

router = APIRouter(tags=["Compliance & Reports"])
compliance_router = APIRouter(prefix="/compliance")
reports_router = APIRouter(prefix="/reports")


@compliance_router.get("/audit-logs", response_model=List[AuditLogOut])
async def get_audit_logs(
    page: int = 1,
    per_page: int = 50,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return paginated audit trail for the current user."""
    result = await db.execute(
        select(AuditLog)
        .where(AuditLog.user_id == current_user.id)
        .order_by(AuditLog.timestamp.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
    )
    return result.scalars().all()

@compliance_router.delete("/audit-logs/{audit_log_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_audit_log(
    audit_log_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete one audit record owned by the current user."""
    result = await db.execute(
        select(AuditLog).where(AuditLog.id == audit_log_id, AuditLog.user_id == current_user.id)
    )
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="Audit record not found")
    await db.delete(record)
    await db.commit()


@compliance_router.delete("/audit-logs")
async def clear_audit_logs(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete all audit records owned by the current user on explicit request."""
    result = await db.execute(select(AuditLog).where(AuditLog.user_id == current_user.id))
    records = result.scalars().all()
    for record in records:
        await db.delete(record)
    await db.commit()
    return {"deleted": len(records)}


@compliance_router.get("/gdpr-export")
async def gdpr_export(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """GDPR data export — all user data from the DB."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    async def rows(model, order=None):
        query = select(model).where(model.user_id == current_user.id)
        if order is not None:
            query = query.order_by(order.desc())
        result = await db.execute(query)
        return result.scalars().all()

    logins = await rows(LoginHistory, LoginHistory.timestamp)
    threats = await rows(Threat, Threat.detected_at)
    alerts = await rows(Alert, Alert.created_at)
    audit_logs = await rows(AuditLog, AuditLog.timestamp)
    simulations = await rows(AttackSimulation, AttackSimulation.created_at)
    sessions = await rows(UserSession, UserSession.created_at)
    devices = await rows(Device, Device.first_seen)
    passkeys = await rows(PasskeyCredential, PasskeyCredential.created_at)
    integrations = await rows(NotificationIntegration, NotificationIntegration.created_at)

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="0EA5E9")
    header_font = Font(color="FFFFFF", bold=True)

    def safe(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            value = value.isoformat()
        elif isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False, default=str)
        elif not isinstance(value, (str, int, float, bool)):
            value = str(value)
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    def add_sheet(title, headers, data):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in data:
            sheet.append([safe(value) for value in row])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, header in enumerate(headers, 1):
            values = [len(str(sheet.cell(row=row, column=index).value or "")) for row in range(1, min(sheet.max_row, 200) + 1)]
            sheet.column_dimensions[get_column_letter(index)].width = min(max([len(header), *values]) + 2, 45)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    generated_at = datetime.now(timezone.utc)
    add_sheet("Profile", ["Field", "Value"], [
        ("Export generated at", generated_at), ("User ID", current_user.id),
        ("Email", current_user.email), ("Username", current_user.username),
        ("Full name", current_user.full_name), ("Role", current_user.role),
        ("Account created", current_user.created_at), ("Last login", current_user.last_login_at),
        ("2FA enabled", current_user.totp_enabled), ("Password breached", current_user.password_breached),
        ("Known breach count", current_user.password_breach_count),
    ])
    add_sheet("Login History", ["Timestamp", "IP", "Country", "City", "Success", "Failure reason", "Device ID", "Simulation"], [
        (x.timestamp, x.ip_address, x.country, x.city, x.success, x.failure_reason, x.device_id, x.is_simulation) for x in logins
    ])
    add_sheet("Threats", ["Detected", "Type", "Severity", "Title", "Description", "Source IP", "Country", "Resolved", "Simulation", "Details"], [
        (x.detected_at, x.threat_type, x.severity, x.title, x.description, x.source_ip, x.source_country, x.is_resolved, x.is_simulation, x.details) for x in threats
    ])
    add_sheet("Alerts", ["Created", "Title", "Message", "Severity", "Read", "Threat ID"], [
        (x.created_at, x.title, x.message, x.severity, x.is_read, x.threat_id) for x in alerts
    ])
    add_sheet("Sessions", ["Created", "Last used", "Expires", "IP", "Device ID", "Active", "Revoked"], [
        (x.created_at, x.last_used_at, x.expires_at, x.ip_address, x.device_id, x.is_active, x.revoked_at) for x in sessions
    ])
    add_sheet("Devices", ["First seen", "Last seen", "Device ID", "Browser", "OS", "Type", "Trusted", "Last IP"], [
        (x.first_seen, x.last_seen, x.device_id, x.browser, x.os, x.device_type, x.is_trusted, x.last_ip) for x in devices
    ])
    add_sheet("Simulations", ["Created", "Type", "Status", "Started", "Ended", "Target", "Summary", "Error"], [
        (x.created_at, x.sim_type, x.status, x.started_at, x.ended_at, x.target_url, x.summary, x.error_message) for x in simulations
    ])
    add_sheet("Audit Log", ["Timestamp", "Action", "Resource", "IP", "Status", "Details"], [
        (x.timestamp, x.action, x.resource, x.ip_address, x.status, x.details) for x in audit_logs
    ])
    add_sheet("Passkeys", ["Name", "Created", "Last used", "Device type", "Backed up", "Transports"], [
        (x.name, x.created_at, x.last_used_at, x.device_type, x.backed_up, x.transports) for x in passkeys
    ])
    add_sheet("Integrations", ["Name", "Type", "Minimum severity", "Active", "Include simulations", "Created", "Last delivery", "Delivery status"], [
        (x.name, x.integration_type, x.minimum_severity, x.is_active, x.include_simulations, x.created_at, x.last_delivery_at, x.last_delivery_status) for x in integrations
    ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"shieldsphere_gdpr_export_{generated_at.date().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@reports_router.post("/generate", response_model=IncidentReportOut)
async def generate_report(
    days: int = 30,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Generate an AI executive security report from real aggregated data."""
    period_end = datetime.now(timezone.utc)
    period_start = period_end - timedelta(days=days)

    # Aggregate real metrics
    result = await db.execute(
        select(func.count(LoginHistory.id)).where(
            LoginHistory.user_id == current_user.id,
            LoginHistory.timestamp >= period_start,
            LoginHistory.is_simulation == False,  # noqa
        )
    )
    total_logins = result.scalar_one() or 0

    result = await db.execute(
        select(func.count(LoginHistory.id)).where(
            LoginHistory.user_id == current_user.id,
            LoginHistory.timestamp >= period_start,
            LoginHistory.success == True,  # noqa
            LoginHistory.is_simulation == False,  # noqa
        )
    )
    successful_logins = result.scalar_one() or 0

    result = await db.execute(
        select(func.count(Threat.id)).where(
            Threat.user_id == current_user.id,
            Threat.detected_at >= period_start,
            Threat.is_simulation == False,  # noqa
        )
    )
    threat_count = result.scalar_one() or 0

    result = await db.execute(
        select(func.count(Threat.id)).where(
            Threat.user_id == current_user.id,
            Threat.detected_at >= period_start,
            Threat.is_resolved == True,  # noqa
            Threat.is_simulation == False,  # noqa
        )
    )
    resolved_threats = result.scalar_one() or 0

    result = await db.execute(
        select(func.count(Threat.id)).where(
            Threat.user_id == current_user.id,
            Threat.detected_at >= period_start,
            Threat.severity.in_(["critical"]),
            Threat.is_simulation == False,  # noqa
        )
    )
    critical_threats = result.scalar_one() or 0

    result = await db.execute(
        select(func.count(AttackSimulation.id)).where(
            AttackSimulation.user_id == current_user.id,
            AttackSimulation.created_at >= period_start,
        )
    )
    simulation_count = result.scalar_one() or 0

    # Top threat types
    result = await db.execute(
        select(Threat.threat_type, func.count(Threat.id).label("count"))
        .where(
            Threat.user_id == current_user.id,
            Threat.detected_at >= period_start,
            Threat.is_simulation == False,  # noqa
        )
        .group_by(Threat.threat_type)
        .order_by(func.count(Threat.id).desc())
        .limit(5)
    )
    top_threat_types = [{"type": r.threat_type, "count": r.count} for r in result.all()]

    # Latest security score
    result = await db.execute(
        select(SecurityScore)
        .where(SecurityScore.user_id == current_user.id)
        .order_by(SecurityScore.computed_at.desc())
        .limit(1)
    )
    score_row = result.scalar_one_or_none()
    security_score = score_row.score if score_row else None

    report_data = {
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "total_logins": total_logins,
        "successful_logins": successful_logins,
        "failed_logins": total_logins - successful_logins,
        "threat_count": threat_count,
        "resolved_threats": resolved_threats,
        "critical_threats": critical_threats,
        "simulation_count": simulation_count,
        "security_score": security_score,
        "top_threat_types": top_threat_types,
    }

    # Generate AI summary
    summary = await generate_executive_report(report_data)

    report = IncidentReport(
        user_id=current_user.id,
        title=f"Security Report — Last {days} Days",
        period_start=period_start,
        period_end=period_end,
        threat_count=threat_count,
        alert_count=0,
        simulation_count=simulation_count,
        executive_summary=summary,
        raw_data=report_data,
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)

    return report

@reports_router.get("/{report_id}/pdf")
async def download_report_pdf(
    report_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download a readable, printable executive security report."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    result = await db.execute(
        select(IncidentReport).where(IncidentReport.id == report_id, IncidentReport.user_id == current_user.id)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    data = report.raw_data or {}
    styles = getSampleStyleSheet()
    navy = colors.HexColor("#102A43")
    blue = colors.HexColor("#0EA5E9")
    pale_blue = colors.HexColor("#E0F2FE")
    body = ParagraphStyle("ShieldBody", parent=styles["BodyText"], textColor=colors.HexColor("#243447"), fontSize=9.5, leading=14, spaceAfter=8)
    heading = ParagraphStyle("ShieldHeading", parent=styles["Heading2"], textColor=navy, fontSize=14, leading=18, spaceBefore=14, spaceAfter=6)
    title = ParagraphStyle("ShieldTitle", parent=styles["Title"], textColor=colors.white, fontSize=23, leading=27, alignment=TA_LEFT, spaceAfter=0)
    sub = ParagraphStyle("ShieldSub", parent=body, textColor=colors.HexColor("#64748B"), fontSize=8.5, alignment=TA_CENTER, spaceAfter=12)
    banner_sub = ParagraphStyle("ShieldBannerSub", parent=body, textColor=colors.HexColor("#BAE6FD"), fontSize=9, leading=12, alignment=TA_LEFT, spaceAfter=0)
    metric_value = ParagraphStyle("ShieldMetricValue", parent=body, textColor=navy, fontSize=19, leading=22, alignment=TA_CENTER, fontName="Helvetica-Bold", spaceAfter=2)
    metric_label = ParagraphStyle("ShieldMetricLabel", parent=body, textColor=colors.HexColor("#475569"), fontSize=7.8, leading=10, alignment=TA_CENTER, spaceAfter=0)

    def value(item, omit_headings: set[str] | None = None):
        """Convert generated report text into clean PDF-safe prose.

        The report model can contain Markdown headings and bold markers. The PDF
        supplies its own typography, so strip those markers and avoid repeating
        a section heading immediately below the section title.
        """
        text = str(item if item is not None else "—")
        cleaned_lines = []
        for line in text.splitlines():
            plain_heading = re.sub(r"^[\s#*_`-]+|[\s:*_`-]+$", "", line).lower()
            if omit_headings and plain_heading in omit_headings:
                continue
            cleaned = re.sub(r"^\s{0,3}#{1,6}\s*", "", line)
            cleaned = cleaned.replace("**", "").replace("__", "")
            cleaned_lines.append(cleaned)
        return escape("\n".join(cleaned_lines).strip() or "—").replace("\n", "<br/>")

    def metrics(rows):
        table = Table([[Paragraph(f"<b>{value(label)}</b>", body), Paragraph(value(item), body)] for label, item in rows], colWidths=[75 * mm, 95 * mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F8FAFC")),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        return table

    def metric_cards(items):
        cards = [[
            [Paragraph(value(item), metric_value), Paragraph(value(label).upper(), metric_label)]
            for label, item in items
        ]]
        table = Table(cards, colWidths=[170 * mm / len(items)] * len(items))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        return table

    def page_chrome(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
        canvas.setLineWidth(0.5)
        canvas.line(doc.leftMargin, 11 * mm, A4[0] - doc.rightMargin, 11 * mm)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.setFont("Helvetica", 7.5)
        canvas.drawString(doc.leftMargin, 7 * mm, "ShieldSphere confidential security report")
        canvas.drawRightString(A4[0] - doc.rightMargin, 7 * mm, f"Page {doc.page}")
        canvas.restoreState()

    executive_summary = report.executive_summary or "No executive summary is available for this report."
    # Some AI responses start with their own Markdown heading. The PDF already
    # has an Executive summary heading, so remove every standalone duplicate.
    executive_summary = re.sub(
        r"(?im)^\s*[#*_`:\- ]*executive\s+(?:security\s+)?summary[#*_`:\- ]*(?:\r?\n|$)",
        "",
        executive_summary,
    ).strip()

    score = data.get("security_score", "Not available")
    story = [
        Table(
            [[Paragraph("ShieldSphere", title), Paragraph("ACCOUNT SECURITY &amp; COMPLIANCE REPORT", banner_sub)]],
            colWidths=[170 * mm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), navy),
                ("LEFTPADDING", (0, 0), (-1, -1), 16), ("RIGHTPADDING", (0, 0), (-1, -1), 16),
                ("TOPPADDING", (0, 0), (-1, -1), 14), ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
            ]),
        ),
        Spacer(1, 8),
        Paragraph(value(report.title), heading),
        Paragraph(f"Generated {value(report.generated_at.strftime('%d %b %Y, %H:%M UTC'))} | Reporting period: {value(report.period_start.strftime('%d %b %Y') if report.period_start else 'Not set')} – {value(report.period_end.strftime('%d %b %Y') if report.period_end else 'Not set')}", sub),
        HRFlowable(width="100%", thickness=1, color=blue, spaceAfter=6),
        Paragraph("Security posture", heading),
        Paragraph("A concise view of account activity and protective coverage during this reporting period.", body),
        metric_cards([
            ("Account security score", f"{score}/100" if score != "Not available" else score),
            ("Threats detected", data.get("threat_count", report.threat_count or 0)),
            ("Failed sign-ins", data.get("failed_logins", 0)),
        ]),
        Spacer(1, 8),
        metrics([
            ("Account security score", f"{data.get('security_score', 'Not available')}/100"),
            ("Login attempts", data.get("total_logins", 0)),
            ("Successful / failed logins", f"{data.get('successful_logins', 0)} / {data.get('failed_logins', 0)}"),
            ("Threats detected / resolved", f"{data.get('threat_count', report.threat_count or 0)} / {data.get('resolved_threats', 0)}"),
            ("Critical threats", data.get("critical_threats", 0)),
            ("Sandbox exercises", data.get("simulation_count", report.simulation_count or 0)),
        ]),
        Paragraph("Executive summary", heading),
        Table([[Paragraph(value(executive_summary or "No executive summary is available for this report."), body)]], colWidths=[170 * mm], style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F0F9FF")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#7DD3FC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ])),
        Paragraph("Key threat trends", heading),
        Paragraph("Threat categories observed in real account activity during the reporting period.", body),
    ]
    top_types = data.get("top_threat_types") or []
    story.append(metrics([("Threat type", "Occurrences")] + [(item.get("type", "Unknown").replace("_", " ").title(), item.get("count", 0)) for item in top_types]) if top_types else Paragraph("No real threats were recorded in the selected period.", body))
    story.extend([
        Paragraph("Recommended actions", heading),
        Paragraph("Prioritized next steps to improve the account security posture.", body),
        Paragraph(value(report.recommendations or "Review unresolved threats, unfamiliar sessions and devices, enforce strong authentication, and continue regular website vulnerability scans."), body),
        Spacer(1, 8),
        Table([[Paragraph("<b>About sandbox exercises:</b> They are intentionally separated from real account activity. They demonstrate detection coverage and do not change the account security score.", body)]], colWidths=[170 * mm], style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ])),
    ])
    output = io.BytesIO()
    SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        title=report.title,
    ).build(story, onFirstPage=page_chrome, onLaterPages=page_chrome)
    pdf_bytes = output.getvalue()
    if not pdf_bytes.startswith(b"%PDF"):
        raise HTTPException(status_code=500, detail="Could not generate a valid PDF report")
    filename = f"shieldsphere_security_report_{report.generated_at.date().isoformat()}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
            "X-Content-Type-Options": "nosniff",
        },
    )


@reports_router.get("", response_model=List[IncidentReportOut])
async def list_reports(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(IncidentReport)
        .where(IncidentReport.user_id == current_user.id)
        .order_by(IncidentReport.generated_at.desc())
        .limit(20)
    )
    return result.scalars().all()


# Combined router
router.include_router(compliance_router)
router.include_router(reports_router)
