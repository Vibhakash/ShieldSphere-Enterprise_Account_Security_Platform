"""Live API smoke tests against a running ShieldSphere backend.

The script creates a disposable account, exercises the selected feature group,
and deletes the account (including its related rows) before exiting.
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
import time
import uuid
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import httpx
import pyotp
from openpyxl import load_workbook


BASE_URL = "http://127.0.0.1:8000/api/v1"


class SmokeRun:
    def __init__(self, group: str) -> None:
        tag = uuid.uuid4().hex[:8]
        self.group = group
        self.email = f"smoke-{tag}@example.com"
        self.username = f"smoke_{tag}"
        self.password = f"Shield!{tag}Aa9#"
        self.token = ""
        self.refresh_token = ""
        self.failures: list[str] = []

    def call(
        self,
        name: str,
        method: str,
        path: str,
        *,
        json: dict | None = None,
        headers: dict[str, str] | None = None,
        timeout: float = 45.0,
        allowed: tuple[int, ...] = (200, 201, 204),
    ) -> httpx.Response | None:
        request_headers = {"Authorization": f"Bearer {self.token}"} if self.token else {}
        request_headers.update(headers or {})
        try:
            response = httpx.request(
                method,
                BASE_URL + path,
                headers=request_headers,
                json=json,
                timeout=timeout,
                follow_redirects=True,
            )
        except Exception as exc:
            self.failures.append(f"{name}: request error: {exc}")
            print(f"FAIL  {name}: request error: {exc}", flush=True)
            return None
        if response.status_code not in allowed:
            detail = response.text.replace("\n", " ")[:240]
            self.failures.append(f"{name}: HTTP {response.status_code}: {detail}")
            print(f"FAIL  {name}: HTTP {response.status_code}: {detail}", flush=True)
        else:
            print(f"PASS  {name}: HTTP {response.status_code}", flush=True)
        return response

    def authenticate(self) -> bool:
        registered = self.call(
            "auth.register",
            "POST",
            "/auth/register",
            json={
                "email": self.email,
                "username": self.username,
                "password": self.password,
                "full_name": "Disposable Smoke User",
            },
        )
        if not registered or registered.status_code != 201:
            return False
        logged_in = self.call(
            "auth.login",
            "POST",
            "/auth/login",
            json={
                "email": self.email,
                "password": self.password,
                "device_fingerprint": "live-smoke-test",
                "user_agent": "ShieldSphere live smoke test",
            },
        )
        if not logged_in or logged_in.status_code != 200:
            return False
        payload = logged_in.json()
        self.token = payload["access_token"]
        self.refresh_token = payload["refresh_token"]
        return True

    def core(self) -> None:
        paths = {
            "auth.me": "/auth/me",
            "dashboard.stats": "/dashboard/stats",
            "dashboard.security_score": "/dashboard/security-score",
            "dashboard.login_history": "/dashboard/login-history",
            "dashboard.login_locations": "/dashboard/login-locations",
            "dashboard.activity_timeline": "/dashboard/activity-timeline",
            "sessions.list": "/sessions",
            "devices.list": "/devices",
            "threats.list": "/threats",
            "alerts.list": "/alerts",
            "blocklist.list": "/ip-blocklist",
            "assessment.url_history": "/assessment/url-scans",
            "assessment.ip_history": "/assessment/ip-reputation",
            "assessment.vulnerability_history": "/assessment/vuln-scans",
            "uba.anomalies": "/uba/anomalies",
            "simulator.types": "/simulator/types",
            "simulator.runs": "/simulator/runs",
            "compliance.audit_logs": "/compliance/audit-logs",
            "reports.list": "/reports",
            "passkeys.list": "/auth/passkeys",
            "integrations.list": "/integrations",
            "secure_account.preview": "/security-actions/containment-preview",
        }
        for name, path in paths.items():
            self.call(name, "GET", path)

        blocked_ip = "203.0.113.25"
        created_block = self.call(
            "blocklist.create",
            "POST",
            "/ip-blocklist",
            json={
                "ip_address": blocked_ip,
                "reason": "Disposable live smoke test",
                "duration": "1h",
            },
        )
        if created_block and created_block.status_code == 201:
            block_data = created_block.json()
            if block_data.get("scope") != "account" or not block_data.get("can_unblock"):
                self.failures.append("blocklist.create: account scope metadata is incorrect")
                print("FAIL  blocklist.create: account scope metadata is incorrect", flush=True)
            self.call(
                "blocklist.enforce_login",
                "POST",
                "/auth/login",
                json={
                    "email": self.email,
                    "password": self.password,
                    "device_fingerprint": "blocked-live-smoke-test",
                    "user_agent": "ShieldSphere blocked IP smoke test",
                },
                headers={"X-Forwarded-For": blocked_ip},
                allowed=(403,),
            )
            self.call("blocklist.list_after_create", "GET", "/ip-blocklist")
            self.call(
                "blocklist.delete",
                "DELETE",
                f"/ip-blocklist/{block_data['id']}",
            )

        # A new account legitimately has no behavior profile until it reaches
        # the minimum login-history threshold.
        self.call("uba.profile", "GET", "/uba/profile", allowed=(200, 404))

        self.call("passkeys.registration_options", "POST", "/auth/passkeys/register/options", json={})
        self.call("passkeys.authentication_options", "POST", "/auth/passkeys/login/options", json={})

        setup = self.call("two_factor.setup", "POST", "/auth/2fa/setup", json={})
        if setup and setup.status_code == 200:
            secret = setup.json()["secret"]
            self.call("two_factor.confirm", "POST", "/auth/2fa/confirm", json={"code": pyotp.TOTP(secret).now()})
            self.call("two_factor.disable", "POST", "/auth/2fa/disable", json={"code": pyotp.TOTP(secret).now()})

        breach = self.call(
            "breach.current_password_sync",
            "POST",
            "/assessment/breach-check",
            json={"password": self.password},
        )
        if breach and breach.status_code == 200:
            data = breach.json()
            if not data.get("is_current_password") or not data.get("account_status_updated"):
                self.failures.append("breach.current_password_sync: account status was not synchronized")
                print("FAIL  breach.current_password_sync: account status was not synchronized", flush=True)

        export = self.call("gdpr.xlsx_export", "GET", "/compliance/gdpr-export")
        if export and export.status_code == 200:
            try:
                workbook = load_workbook(BytesIO(export.content), read_only=True)
                if len(workbook.sheetnames) < 5:
                    raise ValueError("expected multiple GDPR workbook sheets")
                print(f"PASS  gdpr.workbook: {len(workbook.sheetnames)} sheets", flush=True)
            except Exception as exc:
                self.failures.append(f"gdpr.workbook: {exc}")
                print(f"FAIL  gdpr.workbook: {exc}", flush=True)

        self.call("secure_account.apply", "POST", "/security-actions/secure-account", json={})
        self.call("secure_account.current_session_preserved", "GET", "/auth/me")
        self.call(
            "auth.refresh",
            "POST",
            "/auth/refresh",
            json={"refresh_token": self.refresh_token},
        )

    def external(self) -> None:
        self.call(
            "assessment.password_strength",
            "POST",
            "/assessment/password-strength",
            json={"password": self.password},
        )
        self.call("assessment.ip_reputation", "POST", "/assessment/ip-reputation", json={"ip": "8.8.8.8"})
        self.call("assessment.url_scan", "POST", "/assessment/url-scan", json={"url": "https://example.com/"})
        scan = self.call(
            "assessment.vulnerability_scan",
            "POST",
            "/assessment/vuln-scan",
            json={"target_url": "https://example.com/"},
        )
        if scan and scan.status_code in (200, 201):
            scan_id = scan.json()["id"]
            finished = False
            for _ in range(25):
                time.sleep(1)
                history = self.call("assessment.vulnerability_poll", "GET", "/assessment/vuln-scans")
                if not history or history.status_code != 200:
                    break
                item = next((row for row in history.json() if row["id"] == scan_id), None)
                if item and item["status"] in ("completed", "error"):
                    finished = item["status"] == "completed"
                    break
            if not finished:
                self.failures.append("assessment.vulnerability_scan: did not complete successfully")
                print("FAIL  assessment.vulnerability_scan: did not complete successfully", flush=True)

        copilot = self.call(
            "copilot.chat",
            "POST",
            "/copilot/chat",
            json={"message": "Give me three concise steps to secure this account.", "history": []},
            timeout=75.0,
        )
        if copilot and copilot.status_code == 200 and len(copilot.text.strip()) < 20:
            self.failures.append("copilot.chat: response was empty")

        self.call("reports.generate", "POST", "/reports/generate?days=30", json={}, timeout=75.0)

        webhook_url = os.getenv("SMOKE_WEBHOOK_URL")
        integration = self.call(
            "integrations.create_webhook",
            "POST",
            "/integrations",
            json={
                "name": "Disposable smoke webhook",
                "integration_type": "webhook",
                "destination": webhook_url or "https://example.com/security-events",
                "minimum_severity": "low",
                "include_simulations": True,
            },
        )
        if integration and integration.status_code == 201:
            integration_id = integration.json()["id"]
            if webhook_url:
                self.call("integrations.test", "POST", f"/integrations/{integration_id}/test", json={}, timeout=60.0)
            else:
                print("SKIP  integrations.test: set SMOKE_WEBHOOK_URL to a receiver you control", flush=True)
            self.call("integrations.deliveries", "GET", f"/integrations/{integration_id}/deliveries")
            self.call("integrations.toggle", "PATCH", f"/integrations/{integration_id}/toggle", json={})
            self.call("integrations.delete", "DELETE", f"/integrations/{integration_id}")

    def simulator(self) -> None:
        started = self.call(
            "simulator.start",
            "POST",
            "/simulator/run",
            json={
                "sim_type": "brute_force",
                "params": {"attacker_ip": "198.51.100.23", "attempts": 6, "username": "sandbox-user"},
            },
        )
        if not started or started.status_code != 201:
            return
        simulation_id = started.json()["id"]
        terminal = None
        for _ in range(100):
            time.sleep(1)
            current = self.call("simulator.poll", "GET", f"/simulator/runs/{simulation_id}")
            if current and current.status_code == 200:
                terminal = current.json()
                if terminal["status"] in ("completed", "failed"):
                    break
        if not terminal or terminal["status"] != "completed":
            detail = terminal.get("error_message") if terminal else "timed out"
            self.failures.append(f"simulator.completed: {detail}")
            print(f"FAIL  simulator.completed: {detail}", flush=True)
        else:
            print("PASS  simulator.completed", flush=True)
        self.call("simulator.events", "GET", f"/simulator/runs/{simulation_id}/events")
        self.call("simulator.replay", "GET", f"/simulator/runs/{simulation_id}/replay")

    async def cleanup(self) -> None:
        from sqlalchemy import select
        from app.db.models.user import User
        from app.db.session import AsyncSessionLocal

        async with AsyncSessionLocal() as database:
            user = (await database.execute(select(User).where(User.email == self.email))).scalar_one_or_none()
            if user:
                await database.delete(user)
                await database.commit()

    def run(self) -> int:
        try:
            if self.authenticate():
                getattr(self, self.group)()
        finally:
            if sys.platform == "win32":
                asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
            asyncio.run(self.cleanup())
        print(f"SUMMARY group={self.group} failures={len(self.failures)}", flush=True)
        for failure in self.failures:
            print(f" - {failure}", flush=True)
        return 1 if self.failures else 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("group", choices=("core", "external", "simulator"))
    arguments = parser.parse_args()
    raise SystemExit(SmokeRun(arguments.group).run())
